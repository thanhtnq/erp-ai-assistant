import json
import re
from pathlib import Path
from typing import Any


REQUIRED_SECTIONS = {
    "report_catalog", "output_mapping", "filter_mapping",
    "business_synonym", "sample_questions",
}
OPTIONAL_SECTIONS = {
    "table_mapping", "join_relationship", "business_rules",
    "field_mapping", "mandatory_fields", "child_tabs", "sales_cycle",
    "sql_templates", "permissions", "engine_components",
}
ALL_SECTIONS = REQUIRED_SECTIONS | OPTIONAL_SECTIONS

REPORT_REQUIRED = {"report_id", "module", "report_name", "intent_type", "tool_name"}
OUTPUT_REQUIRED = {"report_id", "query_column", "output_name"}
FILTER_REQUIRED = {"report_id", "filter_column"}
SYNONYM_REQUIRED = {"business_term", "technical_term"}
QUESTION_REQUIRED = {"report_id", "user_question"}
RELATIONSHIP_REQUIRED = {"parent_table", "child_table", "join_condition"}
RULE_REQUIRED = {"rule_name", "rule_expression"}
FIELD_REQUIRED = {"table", "field", "business_label"}
MANDATORY_REQUIRED = {"module", "table", "field", "rule"}
CHILD_TAB_REQUIRED = {"module", "parent_document", "child_tab", "child_table"}
CYCLE_REQUIRED = {"module", "sequence", "document_name", "tag_table_usage"}
SQL_TEMPLATE_REQUIRED = {"report_id", "sql_template"}
PERMISSION_REQUIRED = {"module", "rule_name", "rule_expression"}
ENGINE_REQUIRED = {"component", "status", "responsibility"}

ALLOWED_TOOLS = {
    "list_sales_documents", "count_sales_documents", "aggregate_sales_documents", "get_sales_document",
    "list_sales_orders", "count_sales_orders", "aggregate_sales_orders", "get_sales_order",
    "list_sales_invoices", "count_sales_invoices", "aggregate_sales_invoices", "get_sales_invoice",
    "run_query", "run_inventory_query", "get_scm_overview", "run_scm_model",
}
BLOCKED_SQL_PATTERN = re.compile(r"\b(insert|update|delete|drop|alter|truncate|merge|exec)\b", re.IGNORECASE)


class SemanticValidationError(ValueError):
    pass


def _norm_key(value: str) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _clean_row(row: dict[str, Any]) -> dict[str, Any]:
    cleaned = {}
    for key, value in row.items():
        nk = _norm_key(key)
        if not nk:
            continue
        if isinstance(value, str):
            value = value.strip()
        cleaned[nk] = value
    return cleaned


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "required"}


def _as_list(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("["):
            try:
                loaded = json.loads(text)
                if isinstance(loaded, list):
                    return [str(v).strip() for v in loaded if str(v).strip()]
            except Exception:
                pass
        return [part.strip() for part in text.split(",") if part.strip()]
    return [str(value).strip()]


def _as_dict(value: Any) -> dict[str, Any]:
    if value is None or value == "":
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return {}
        try:
            loaded = json.loads(text)
            if isinstance(loaded, dict):
                return loaded
        except Exception:
            pass
        pairs = {}
        for part in text.split(","):
            if "=" in part:
                key, val = part.split("=", 1)
                pairs[key.strip()] = val.strip()
        return pairs
    return {}


def _require(row: dict[str, Any], fields: set[str], section: str, index: int) -> None:
    missing = [field for field in sorted(fields) if row.get(field) in (None, "")]
    if missing:
        raise SemanticValidationError(f"{section} row {index}: missing {', '.join(missing)}")


def load_metadata_file(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    path = Path(path)
    if not path.exists():
        raise SemanticValidationError(f"File not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    raise SemanticValidationError("Only .json and .xlsx metadata files are supported")


def _load_json(path: Path) -> dict[str, list[dict[str, Any]]]:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise SemanticValidationError(f"Invalid JSON: {exc}") from exc
    if not isinstance(raw, dict):
        raise SemanticValidationError("Metadata JSON must be an object")
    data = {}
    for key, rows in raw.items():
        section = _norm_key(key)
        if section not in ALL_SECTIONS:
            continue
        if not isinstance(rows, list):
            raise SemanticValidationError(f"{section} must be a list")
        data[section] = [_clean_row(row) for row in rows if isinstance(row, dict)]
    return data


def _load_xlsx(path: Path) -> dict[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SemanticValidationError("openpyxl is required to ingest .xlsx metadata") from exc
    wb = load_workbook(path, data_only=True)
    data: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        section = _norm_key(sheet_name)
        if section not in ALL_SECTIONS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            data[section] = []
            continue
        headers = [_norm_key(h) for h in rows[0]]
        section_rows = []
        for raw in rows[1:]:
            if not raw or all(v is None or v == "" for v in raw):
                continue
            section_rows.append(_clean_row(dict(zip(headers, raw))))
        data[section] = section_rows
    return data


def validate_metadata(data: dict[str, list[dict[str, Any]]], expected_module: str | None = None) -> dict[str, list[dict[str, Any]]]:
    missing_sections = sorted(section for section in REQUIRED_SECTIONS if section not in data)
    if missing_sections:
        raise SemanticValidationError(f"Missing required section(s): {', '.join(missing_sections)}")

    normalized = {section: list(data.get(section, [])) for section in ALL_SECTIONS}
    if not normalized["report_catalog"]:
        raise SemanticValidationError("report_catalog must contain at least one report")

    report_ids = set()
    for idx, row in enumerate(normalized["report_catalog"], 1):
        _require(row, REPORT_REQUIRED, "report_catalog", idx)
        row["report_id"] = str(row["report_id"]).strip()
        row["module"] = str(row["module"]).strip().lower()
        if expected_module and row["module"] != expected_module:
            raise SemanticValidationError(
                f"report_catalog row {idx}: module {row['module']} does not match upload module {expected_module}"
            )
        if row["report_id"] in report_ids:
            raise SemanticValidationError(f"Duplicate report_id in file: {row['report_id']}")
        report_ids.add(row["report_id"])
        if row["tool_name"] not in ALLOWED_TOOLS:
            raise SemanticValidationError(f"report_catalog row {idx}: unknown tool_name {row['tool_name']}")
        row["default_filters"] = _as_dict(row.get("default_filters"))
        row["required_filters"] = _as_list(row.get("required_filters"))

    for section, required in (
        ("output_mapping", OUTPUT_REQUIRED),
        ("filter_mapping", FILTER_REQUIRED),
        ("sample_questions", QUESTION_REQUIRED),
    ):
        for idx, row in enumerate(normalized[section], 1):
            _require(row, required, section, idx)
            row["report_id"] = str(row["report_id"]).strip()
            if row["report_id"] not in report_ids:
                raise SemanticValidationError(f"{section} row {idx}: unknown report_id {row['report_id']}")

    for idx, row in enumerate(normalized["filter_mapping"], 1):
        row["required"] = _as_bool(row.get("required"))

    for idx, row in enumerate(normalized["business_synonym"], 1):
        _require(row, SYNONYM_REQUIRED, "business_synonym", idx)

    for idx, row in enumerate(normalized["join_relationship"], 1):
        _require(row, RELATIONSHIP_REQUIRED, "join_relationship", idx)

    for idx, row in enumerate(normalized["business_rules"], 1):
        _require(row, RULE_REQUIRED, "business_rules", idx)
        if row.get("report_id") and row["report_id"] not in report_ids:
            raise SemanticValidationError(f"business_rules row {idx}: unknown report_id {row['report_id']}")

    for idx, row in enumerate(normalized["field_mapping"], 1):
        _require(row, FIELD_REQUIRED, "field_mapping", idx)

    for idx, row in enumerate(normalized["mandatory_fields"], 1):
        _require(row, MANDATORY_REQUIRED, "mandatory_fields", idx)
        row["module"] = str(row["module"]).strip().lower()
        if expected_module and row["module"] != expected_module:
            raise SemanticValidationError(
                f"mandatory_fields row {idx}: module {row['module']} does not match upload module {expected_module}"
            )

    for idx, row in enumerate(normalized["child_tabs"], 1):
        _require(row, CHILD_TAB_REQUIRED, "child_tabs", idx)
        row["module"] = str(row["module"]).strip().lower()
        if expected_module and row["module"] != expected_module:
            raise SemanticValidationError(
                f"child_tabs row {idx}: module {row['module']} does not match upload module {expected_module}"
            )

    for idx, row in enumerate(normalized["sales_cycle"], 1):
        _require(row, CYCLE_REQUIRED, "sales_cycle", idx)
        row["module"] = str(row["module"]).strip().lower()
        if expected_module and row["module"] != expected_module:
            raise SemanticValidationError(
                f"sales_cycle row {idx}: module {row['module']} does not match upload module {expected_module}"
            )
        try:
            row["sequence"] = int(row["sequence"])
        except Exception as exc:
            raise SemanticValidationError(f"sales_cycle row {idx}: sequence must be a number") from exc

    for idx, row in enumerate(normalized["sql_templates"], 1):
        _require(row, SQL_TEMPLATE_REQUIRED, "sql_templates", idx)
        row["report_id"] = str(row["report_id"]).strip()
        if row["report_id"] not in report_ids:
            raise SemanticValidationError(f"sql_templates row {idx}: unknown report_id {row['report_id']}")
        sql = str(row.get("sql_template") or "").strip()
        if not sql.lower().startswith("select"):
            raise SemanticValidationError(f"sql_templates row {idx}: only SELECT templates are allowed")
        if BLOCKED_SQL_PATTERN.search(sql):
            raise SemanticValidationError(f"sql_templates row {idx}: unsafe SQL keyword found")

    for idx, row in enumerate(normalized["permissions"], 1):
        _require(row, PERMISSION_REQUIRED, "permissions", idx)
        row["module"] = str(row["module"]).strip().lower()
        if expected_module and row["module"] != expected_module:
            raise SemanticValidationError(
                f"permissions row {idx}: module {row['module']} does not match upload module {expected_module}"
            )

    for idx, row in enumerate(normalized["engine_components"], 1):
        _require(row, ENGINE_REQUIRED, "engine_components", idx)

    return normalized


def load_and_validate(path: str | Path, expected_module: str | None = None) -> dict[str, list[dict[str, Any]]]:
    return validate_metadata(load_metadata_file(path), expected_module=expected_module)
